"""要求清单 Excel 导出。

本期仅导出；import_requirements 为二期预留接口。
导出的 xlsx 中分类/置信度/状态列带数据校验下拉，保证在 Excel 内编辑时值合法，便于将来导回。
"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CATEGORIES = ["资质门槛", "评分项", "废标项", "格式要求", "时间节点", "其他"]
CONFIDENCES = ["高", "中", "低"]
STATUSES = ["已响应", "待响应", "需关注"]

HEADERS = ["序号", "分类", "内容", "来源章节", "置信度", "状态"]
COLUMN_WIDTHS = [6, 12, 60, 16, 10, 10]

HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(bold=True, color="FFFFFF")

STATUS_FILLS = {
    "已响应": PatternFill("solid", fgColor="DCFCE7"),
    "待响应": PatternFill("solid", fgColor="FEF3C7"),
    "需关注": PatternFill("solid", fgColor="FEE2E2"),
}


def export_requirements(project: dict, requirements: list, dest_path: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "要求清单"

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, req in enumerate(requirements, start=1):
        ws.append([
            i,
            req.get("category", ""),
            req.get("content", ""),
            req.get("source", ""),
            req.get("confidence", ""),
            req.get("status", ""),
        ])
        status = req.get("status", "")
        if status in STATUS_FILLS:
            ws.cell(row=i + 1, column=6).fill = STATUS_FILLS[status]
        ws.cell(row=i + 1, column=3).alignment = Alignment(wrap_text=True, vertical="top")

    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    last_row = max(len(requirements) + 1, 2)
    ws.auto_filter.ref = f"A1:F{last_row}"

    def add_validation(col_letter, options):
        dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")
        ws.add_data_validation(dv)

    add_validation("B", CATEGORIES)
    add_validation("E", CONFIDENCES)
    add_validation("F", STATUSES)

    info = wb.create_sheet("说明")
    info_rows = [
        ["项目名称", project.get("name", "")],
        ["招标单位", project.get("client", "")],
        ["投标日期", project.get("bid_date", "")],
        ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [],
        ["说明", "本文件为导出副本，在 Excel 中的修改不会同步回投标助手程序。"],
        ["分类可选值", "、".join(CATEGORIES)],
        ["置信度可选值", "、".join(CONFIDENCES)],
        ["状态可选值", "、".join(STATUSES)],
    ]
    for row in info_rows:
        info.append(row)
    info.column_dimensions["A"].width = 14
    info.column_dimensions["B"].width = 60

    dest = Path(dest_path)
    wb.save(dest)
    return dest


def import_requirements(*args, **kwargs):
    raise NotImplementedError("Excel 导回功能将在第二阶段实现；导出文件已含数据校验下拉以便后续导回。")
