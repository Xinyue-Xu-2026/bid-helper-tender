from pathlib import Path

import pytest
from openpyxl import load_workbook

from bidhelper.excel_exporter import export_requirements, import_requirements

PROJECT = {"name": "测试项目", "client": "某单位", "bid_date": "2026-08-15"}
REQUIREMENTS = [
    {"category": "废标项", "content": "投标文件必须签字盖章", "source": "第四章", "confidence": "高", "status": "需关注"},
    {"category": "格式要求", "content": "统一用A4规格幅面打印、装订成册并编制目录", "source": "2.5", "confidence": "中", "status": "待响应"},
    {"category": "时间节点", "content": "2026年8月13日9时00分", "source": "投标截止时间", "confidence": "高", "status": "已响应"},
]


@pytest.fixture
def dest(tmp_path):
    return str(tmp_path / "out.xlsx")


def test_export_creates_file(dest):
    result = export_requirements(PROJECT, REQUIREMENTS, dest)
    assert Path(result).exists()


def test_export_sheet_content(dest):
    export_requirements(PROJECT, REQUIREMENTS, dest)
    ws = load_workbook(dest)["要求清单"]
    assert [c.value for c in ws[1]] == ["序号", "分类", "内容", "来源章节", "置信度", "状态"]
    assert ws.max_row == 4
    row2 = [ws.cell(row=2, column=c).value for c in range(1, 7)]
    assert row2 == [1, "废标项", "投标文件必须签字盖章", "第四章", "高", "需关注"]


def test_freeze_and_filter(dest):
    export_requirements(PROJECT, REQUIREMENTS, dest)
    ws = load_workbook(dest)["要求清单"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:F4"


def test_data_validations(dest):
    export_requirements(PROJECT, REQUIREMENTS, dest)
    ws = load_workbook(dest)["要求清单"]
    formulas = {dv.formula1 for dv in ws.data_validations.dataValidation}
    assert any("废标项" in f and "资质门槛" in f for f in formulas)
    assert any("已响应" in f and "需关注" in f for f in formulas)
    assert any("高" in f and "低" in f for f in formulas)


def test_status_fill_colors(dest):
    export_requirements(PROJECT, REQUIREMENTS, dest)
    ws = load_workbook(dest)["要求清单"]
    assert ws.cell(row=2, column=6).fill.fgColor.rgb.endswith("FEE2E2")   # 需关注=浅红
    assert ws.cell(row=3, column=6).fill.fgColor.rgb.endswith("FEF3C7")   # 待响应=浅橙
    assert ws.cell(row=4, column=6).fill.fgColor.rgb.endswith("DCFCE7")   # 已响应=浅绿


def test_info_sheet(dest):
    export_requirements(PROJECT, REQUIREMENTS, dest)
    info = load_workbook(dest)["说明"]
    assert info.cell(row=1, column=2).value == "测试项目"
    assert info.cell(row=4, column=1).value == "导出时间"


def test_import_not_implemented():
    with pytest.raises(NotImplementedError):
        import_requirements("x")
